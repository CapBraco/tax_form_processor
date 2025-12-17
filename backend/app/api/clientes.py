"""
Clientes API Endpoints - Phase 3: User Data Isolation
✅ All queries filtered by user_id
✅ Complete Excel/PDF export with proper styling
✅ Works with NULL periodo_anio documents
✅ Year validation
✅ FIXED: Only use Form 104 fields that actually exist in the database
"""

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from typing import List, Dict, Optional
from pydantic import BaseModel

import requests
from PIL import Image as PILImage

from app.core.database import get_db
from app.models.base import Document, Form103Totals, Form104Data, FormTypeEnum
from app.core.security import get_current_user
from app.models.base import User

from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, Image, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

router = APIRouter(prefix="/clientes", tags=["clientes"])

# Spanish month → number mapping
MONTH_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3,
    "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9,
    "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
}

# Pydantic models
class ClientSummary(BaseModel):
    razon_social: str
    document_count: int
    first_year: str
    last_year: str

class FormInfo(BaseModel):
    id: int
    filename: str
    uploaded_at: str
    identificacion_ruc: Optional[str]

class MonthData(BaseModel):
    month: int
    periodo_fiscal: Optional[str]
    forms: Dict[str, Optional[FormInfo]]

class YearData(BaseModel):
    year: str
    months: List[MonthData]

class ClientDocuments(BaseModel):
    razon_social: str
    years: List[YearData]

class PDFBranding(BaseModel):
    company_name: str
    logo_url: Optional[str] = None
    primary_color: Optional[str] = "#1a73e8"
    secondary_color: Optional[str] = "#34a853"
    footer_text: str


def is_valid_year(year: str) -> bool:
    """Check if year is valid (not None, 'Unknown', empty, etc.)"""
    if not year or year.upper() in ('UNKNOWN', 'N/A', ''):
        return False
    try:
        year_int = int(year)
        return 1900 <= year_int <= 2100
    except (ValueError, TypeError):
        return False


# ------------------------------
# List all clients
# ------------------------------
@router.get("/", response_model=List[ClientSummary])
async def get_all_clients(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all clients from current user's documents"""
    query = select(
        Document.razon_social,
        func.count(Document.id).label('document_count'),
        func.min(Document.periodo_anio).label('first_year'),
        func.max(Document.periodo_anio).label('last_year')
    ).where(
        and_(
            Document.razon_social.isnot(None),
            Document.user_id == current_user.id
        )
    ).group_by(Document.razon_social).order_by(Document.razon_social)

    result = await db.execute(query)
    clients = result.all()

    return [
        ClientSummary(
            razon_social=c.razon_social,
            document_count=c.document_count,
            first_year=c.first_year or "N/A",
            last_year=c.last_year or "N/A"
        ) for c in clients
    ]


# ------------------------------
# Documents grouped by year/month
# ------------------------------
@router.get("/{razon_social}")
async def get_client_documents(
    razon_social: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get documents for a specific client"""
    query = select(Document).where(
        and_(
            Document.razon_social == razon_social,
            Document.user_id == current_user.id
        )
    ).order_by(Document.periodo_anio.desc(), Document.id.asc())
    
    result = await db.execute(query)
    documents = result.scalars().all()
    
    if not documents:
        raise HTTPException(
            status_code=404,
            detail=f"No documents found for client '{razon_social}' in your account"
        )

    organized = {}
    documents_with_period = 0
    
    for doc in documents:
        if not doc.periodo_anio:
            continue
        
        documents_with_period += 1
        year = doc.periodo_anio
        
        if year not in organized:
            organized[year] = {'year': year, 'months': {}}

        month_num = 0
        if doc.periodo_mes:
            month_num = MONTH_MAP.get(doc.periodo_mes.upper().strip(), 0)
        periodo_fiscal = f"{doc.periodo_mes} {doc.periodo_anio}" if doc.periodo_mes else None

        if month_num not in organized[year]['months']:
            organized[year]['months'][month_num] = {
                'month': month_num, 
                'periodo_fiscal': periodo_fiscal, 
                'forms': {'form_103': None, 'form_104': None}
            }

        form_key = 'form_103' if doc.form_type == FormTypeEnum.FORM_103 else 'form_104'
        organized[year]['months'][month_num]['forms'][form_key] = {
            'id': doc.id,
            'filename': doc.original_filename,
            'uploaded_at': doc.uploaded_at.isoformat() if doc.uploaded_at else None,
            'identificacion_ruc': doc.identificacion_ruc
        }

    if documents_with_period == 0:
        raise HTTPException(
            status_code=404,
            detail=f"Client found with {len(documents)} documents, but none have valid period information. Please reprocess documents."
        )

    result_years = []
    for year_data in organized.values():
        months_list = sorted(year_data['months'].values(), key=lambda x: x['month'])
        year_data['months'] = months_list
        result_years.append(year_data)
    result_years.sort(key=lambda x: x['year'], reverse=True)

    return {'razon_social': razon_social, 'years': result_years}


# ------------------------------
# Yearly summary - ✅ FIXED with getattr()
# ------------------------------
@router.get("/{razon_social}/yearly-summary/{year}")
async def get_yearly_summary(
    razon_social: str, 
    year: str, 
    exclude_months: Optional[str] = None, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get yearly summary for a client"""
    if not is_valid_year(year):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid year parameter: '{year}'. Year must be a valid 4-digit year."
        )
    
    excluded = set(int(m.strip()) for m in exclude_months.split(',')) if exclude_months else set()

    query_103 = select(Document).where(
        and_(
            Document.razon_social == razon_social, 
            Document.form_type == FormTypeEnum.FORM_103, 
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    )
    if excluded:
        query_103 = query_103.where(Document.periodo_mes_numero.notin_(excluded))
    docs_103 = (await db.execute(query_103)).scalars().all()

    query_104 = select(Document).where(
        and_(
            Document.razon_social == razon_social, 
            Document.form_type == FormTypeEnum.FORM_104, 
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    )
    if excluded:
        query_104 = query_104.where(Document.periodo_mes_numero.notin_(excluded))
    docs_104 = (await db.execute(query_104)).scalars().all()

    summary_103 = {
        'subtotal_operaciones_pais': 0.0,
        'total_retencion': 0.0,
        'total_impuesto_pagar': 0.0,
        'total_pagado': 0.0,
        'monthly_details': []
    }
    
    # ✅ FIXED: Include ALL Form 104 fields that ACTUALLY exist in the database
    summary_104 = {
        # Basic fields
        'total_ventas_neto': 0.0,
        'total_impuesto_generado': 0.0,
        'total_adquisiciones': 0.0,
        'credito_tributario_aplicable': 0.0,
        'total_impuesto_retenido': 0.0,
        'total_pagado': 0.0,
        
        # Calculated fields that EXIST in database
        'impuesto_causado': 0.0,
        'retenciones_efectuadas': 0.0,
        'subtotal_a_pagar': 0.0,
        'total_impuesto_pagar_retencion': 0.0,
        'total_consolidado_iva': 0.0,
        'total_impuesto_a_pagar': 0.0,
        'interes_mora': 0.0,
        'multa': 0.0,
        
        'monthly_details': []
    }

    # --- Form 103 processing ---
    for doc in docs_103:
        tot = (await db.execute(select(Form103Totals).where(Form103Totals.document_id == doc.id))).scalar_one_or_none()
        if not tot:
            continue
        periodo_fiscal = f"{doc.periodo_mes} {doc.periodo_anio}" if doc.periodo_mes else None
        md = {
            'month': doc.periodo_mes_numero,
            'periodo_fiscal': periodo_fiscal,
            'subtotal_operaciones_pais': tot.subtotal_operaciones_pais or 0.0,
            'total_retencion': tot.total_retencion or 0.0,
            'total_impuesto_pagar': tot.total_impuesto_pagar or 0.0,
            'total_pagado': tot.total_pagado or 0.0
        }
        for k in ['subtotal_operaciones_pais', 'total_retencion', 'total_impuesto_pagar', 'total_pagado']:
            summary_103[k] += md[k]
        summary_103['monthly_details'].append(md)

    # --- Form 104 processing - ✅ FIXED with getattr() for safety ---
    for doc in docs_104:
        data = (await db.execute(select(Form104Data).where(Form104Data.document_id == doc.id))).scalar_one_or_none()
        
        if not data:
            continue
            
        periodo_fiscal = f"{doc.periodo_mes} {doc.periodo_anio}" if doc.periodo_mes else None
        
        # ✅ Use getattr() with 0 default - safe for all fields
        md = {
            'month': doc.periodo_mes_numero,
            'periodo_fiscal': periodo_fiscal,
            'total_ventas_neto': getattr(data, 'total_ventas_neto', 0) or 0.0,
            'total_impuesto_generado': getattr(data, 'total_impuesto_generado', 0) or 0.0,
            'total_adquisiciones': getattr(data, 'total_adquisiciones', 0) or 0.0,
            'credito_tributario_aplicable': getattr(data, 'credito_tributario_aplicable', 0) or 0.0,
            'total_impuesto_retenido': getattr(data, 'total_impuesto_retenido', 0) or 0.0,
            'total_pagado': getattr(data, 'total_pagado', 0) or 0.0,
            # Add calculated fields with getattr for safety
            'impuesto_causado': getattr(data, 'impuesto_causado', 0) or 0.0,
            'retenciones_efectuadas': getattr(data, 'retenciones_efectuadas', 0) or 0.0,
            'subtotal_a_pagar': getattr(data, 'subtotal_a_pagar', 0) or 0.0,
            'total_impuesto_pagar_retencion': getattr(data, 'total_impuesto_pagar_retencion', 0) or 0.0,
            'total_consolidado_iva': getattr(data, 'total_consolidado_iva', 0) or 0.0,
            'total_impuesto_a_pagar': getattr(data, 'total_impuesto_a_pagar', 0) or 0.0,
            'interes_mora': getattr(data, 'interes_mora', 0) or 0.0,
            'multa': getattr(data, 'multa', 0) or 0.0,
        }
        
        # Sum all fields into summary
        for key, value in md.items():
            if key not in ['month', 'periodo_fiscal']:
                summary_104[key] += value
        
        summary_104['monthly_details'].append(md)

    all_months = set(range(1, 13))
    present_103 = set(d.periodo_mes_numero for d in docs_103 if d.periodo_mes_numero)
    present_104 = set(d.periodo_mes_numero for d in docs_104 if d.periodo_mes_numero)
    missing_103 = sorted(all_months - present_103 - excluded)
    missing_104 = sorted(all_months - present_104 - excluded)

    return {
        'razon_social': razon_social,
        'year': year,
        'form_103_summary': summary_103,
        'form_104_summary': summary_104,
        'missing_months': {'form_103': missing_103, 'form_104': missing_104},
        'excluded_months': sorted(excluded)
    }


# ------------------------------
# Validation
# ------------------------------
@router.get("/{razon_social}/validation/{year}")
async def validate_year_completeness(
    razon_social: str, 
    year: str, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not is_valid_year(year):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid year parameter: '{year}'. Year must be a valid 4-digit year."
        )
    
    docs_103 = (await db.execute(
        select(Document).where(
            and_(
                Document.razon_social == razon_social,
                Document.form_type == FormTypeEnum.FORM_103,
                Document.periodo_anio == year,
                Document.user_id == current_user.id
            )
        )
    )).scalars().all()
    
    docs_104 = (await db.execute(
        select(Document).where(
            and_(
                Document.razon_social == razon_social,
                Document.form_type == FormTypeEnum.FORM_104,
                Document.periodo_anio == year,
                Document.user_id == current_user.id
            )
        )
    )).scalars().all()

    present_103 = {d.periodo_mes_numero: d for d in docs_103 if d.periodo_mes_numero}
    present_104 = {d.periodo_mes_numero: d for d in docs_104 if d.periodo_mes_numero}

    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    validation = []
    for m in range(1, 13):
        validation.append({
            'month': m,
            'month_name': month_names[m],
            'has_form_103': m in present_103,
            'has_form_104': m in present_104,
            'form_103_id': present_103[m].id if m in present_103 else None,
            'form_104_id': present_104[m].id if m in present_104 else None,
            'is_complete': m in present_103 and m in present_104
        })
    
    complete_months = sum(1 for v in validation if v['is_complete'])
    return {
        'razon_social': razon_social,
        'year': year,
        'complete_months': complete_months,
        'total_months': 12,
        'is_fully_complete': complete_months == 12,
        'validation_details': validation
    }


# ------------------------------
# Export to Excel
# ------------------------------
@router.post("/{razon_social}/export-excel/{year}")
async def export_yearly_excel(
    razon_social: str,
    year: str,
    exclude_months: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export yearly summary to Excel file
    ✅ Complete styling with borders and formatting
    ✅ User data isolation
    """
    if not is_valid_year(year):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid year parameter: '{year}'. Cannot export data for invalid year."
        )
    
    excluded = set(int(m.strip()) for m in exclude_months.split(',')) if exclude_months else set()

    # Fetch Form 103 documents
    query_103 = select(Document).where(
        and_(
            Document.razon_social == razon_social,
            Document.form_type == FormTypeEnum.FORM_103,
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    ).order_by(Document.periodo_mes_numero.asc())
    if excluded:
        query_103 = query_103.where(Document.periodo_mes_numero.notin_(excluded))
    docs_103 = (await db.execute(query_103)).scalars().all()

    # Fetch Form 104 documents
    query_104 = select(Document).where(
        and_(
            Document.razon_social == razon_social,
            Document.form_type == FormTypeEnum.FORM_104,
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    ).order_by(Document.periodo_mes_numero.asc())
    if excluded:
        query_104 = query_104.where(Document.periodo_mes_numero.notin_(excluded))
    docs_104 = (await db.execute(query_104)).scalars().all()

    wb = openpyxl.Workbook()
    
    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    month_names = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    # --- Form 103 Sheet ---
    ws_103 = wb.active
    ws_103.title = "Form 103 - Retenciones"
    
    # Title
    ws_103['A1'] = f"Resumen Anual {year} - Form 103 (Retenciones en la Fuente)"
    ws_103['A1'].font = Font(size=14, bold=True)
    ws_103['A2'] = razon_social
    ws_103['A2'].font = Font(size=11, bold=True)
    ws_103['A3'] = f"Usuario: {current_user.email}"
    ws_103['A3'].font = Font(size=9, italic=True)
    
    # Column headers
    headers_103 = ['Mes', 'Período', 'Subtotal Op. País', 'Total Retención', 'Total Impuesto', 'Total Pagado']
    for col_num, header in enumerate(headers_103, 1):
        cell = ws_103.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = 6
    total_subtotal = 0
    total_retencion = 0
    total_impuesto = 0
    total_pagado = 0
    
    for doc in docs_103:
        tot = (await db.execute(select(Form103Totals).where(Form103Totals.document_id == doc.id))).scalar_one_or_none()
        if not tot:
            continue
        
        periodo_fiscal = f"{doc.periodo_mes} {doc.periodo_anio}" if doc.periodo_mes else "N/A"
        
        ws_103.cell(row=row_num, column=1).value = month_names[doc.periodo_mes_numero] if doc.periodo_mes_numero else "N/A"
        ws_103.cell(row=row_num, column=2).value = periodo_fiscal
        ws_103.cell(row=row_num, column=3).value = tot.subtotal_operaciones_pais or 0.0
        ws_103.cell(row=row_num, column=4).value = tot.total_retencion or 0.0
        ws_103.cell(row=row_num, column=5).value = tot.total_impuesto_pagar or 0.0
        ws_103.cell(row=row_num, column=6).value = tot.total_pagado or 0.0
        
        # Format and borders
        for col in range(1, 7):
            cell = ws_103.cell(row=row_num, column=col)
            cell.border = border
            if col >= 3:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        
        total_subtotal += tot.subtotal_operaciones_pais or 0.0
        total_retencion += tot.total_retencion or 0.0
        total_impuesto += tot.total_impuesto_pagar or 0.0
        total_pagado += tot.total_pagado or 0.0
        
        row_num += 1
    
    # Total row
    ws_103.cell(row=row_num, column=1).value = "TOTAL ANUAL"
    ws_103.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws_103.cell(row=row_num, column=3).value = total_subtotal
    ws_103.cell(row=row_num, column=4).value = total_retencion
    ws_103.cell(row=row_num, column=5).value = total_impuesto
    ws_103.cell(row=row_num, column=6).value = total_pagado
    
    for col in range(1, 7):
        cell = ws_103.cell(row=row_num, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col >= 3:
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
    
    # Column widths
    ws_103.column_dimensions['A'].width = 12
    ws_103.column_dimensions['B'].width = 15
    ws_103.column_dimensions['C'].width = 18
    ws_103.column_dimensions['D'].width = 18
    ws_103.column_dimensions['E'].width = 18
    ws_103.column_dimensions['F'].width = 18
    
    # --- Form 104 Sheet ---
    ws_104 = wb.create_sheet("Form 104 - IVA")
    
    # Title
    ws_104['A1'] = f"Resumen Anual {year} - Form 104 (IVA)"
    ws_104['A1'].font = Font(size=14, bold=True)
    ws_104['A2'] = razon_social
    ws_104['A2'].font = Font(size=11, bold=True)
    ws_104['A3'] = f"Usuario: {current_user.email}"
    ws_104['A3'].font = Font(size=9, italic=True)
    
    # Monthly detail section
    ws_104['A5'] = "DETALLE MENSUAL"
    ws_104['A5'].font = Font(size=12, bold=True)
    
    # Headers
    headers_104 = ['Mes', 'Período', 'Ventas Neto', 'Imp. Generado', 'Adquisiciones', 'Créd. Trib.', 'Imp. Retenido', 'Total Pagado']
    for col_num, header in enumerate(headers_104, 1):
        cell = ws_104.cell(row=6, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Data rows
    row_num = 7
    total_ventas = 0
    total_impuesto_gen = 0
    total_adq = 0
    total_cred = 0
    total_ret = 0
    total_pag = 0
    
    for doc in docs_104:
        data = (await db.execute(select(Form104Data).where(Form104Data.document_id == doc.id))).scalar_one_or_none()
        
        if not data:
            continue
        
        periodo_fiscal = f"{doc.periodo_mes} {doc.periodo_anio}" if doc.periodo_mes else "N/A"
        
        ws_104.cell(row=row_num, column=1).value = month_names[doc.periodo_mes_numero] if doc.periodo_mes_numero else "N/A"
        ws_104.cell(row=row_num, column=2).value = periodo_fiscal
        ws_104.cell(row=row_num, column=3).value = getattr(data, 'total_ventas_neto', 0) or 0.0
        ws_104.cell(row=row_num, column=4).value = getattr(data, 'total_impuesto_generado', 0) or 0.0
        ws_104.cell(row=row_num, column=5).value = getattr(data, 'total_adquisiciones', 0) or 0.0
        ws_104.cell(row=row_num, column=6).value = getattr(data, 'credito_tributario_aplicable', 0) or 0.0
        ws_104.cell(row=row_num, column=7).value = getattr(data, 'total_impuesto_retenido', 0) or 0.0
        ws_104.cell(row=row_num, column=8).value = getattr(data, 'total_pagado', 0) or 0.0
        
        # Format and borders
        for col in range(1, 9):
            cell = ws_104.cell(row=row_num, column=col)
            cell.border = border
            if col >= 3:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        
        total_ventas += getattr(data, 'total_ventas_neto', 0) or 0.0
        total_impuesto_gen += getattr(data, 'total_impuesto_generado', 0) or 0.0
        total_adq += getattr(data, 'total_adquisiciones', 0) or 0.0
        total_cred += getattr(data, 'credito_tributario_aplicable', 0) or 0.0
        total_ret += getattr(data, 'total_impuesto_retenido', 0) or 0.0
        total_pag += getattr(data, 'total_pagado', 0) or 0.0
        
        row_num += 1
    
    # Total row
    ws_104.cell(row=row_num, column=1).value = "TOTAL ANUAL"
    ws_104.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    ws_104.cell(row=row_num, column=3).value = total_ventas
    ws_104.cell(row=row_num, column=4).value = total_impuesto_gen
    ws_104.cell(row=row_num, column=5).value = total_adq
    ws_104.cell(row=row_num, column=6).value = total_cred
    ws_104.cell(row=row_num, column=7).value = total_ret
    ws_104.cell(row=row_num, column=8).value = total_pag
    
    for col in range(1, 9):
        cell = ws_104.cell(row=row_num, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border
        if col >= 3:
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
    
    # Summary section
    row_num += 3
    ws_104.cell(row=row_num, column=1).value = "RESUMEN ANUAL"
    ws_104.cell(row=row_num, column=1).font = Font(size=12, bold=True)
    
    row_num += 1
    summary_items = [
        ('Total Ventas Neto', total_ventas),
        ('Total Impuesto Generado', total_impuesto_gen),
        ('Total Adquisiciones', total_adq),
        ('Total Crédito Tributario', total_cred),
        ('Total Impuesto Retenido', total_ret),
        ('TOTAL PAGADO', total_pag)
    ]
    
    for label, value in summary_items:
        ws_104.cell(row=row_num, column=1).value = label
        ws_104.cell(row=row_num, column=1).font = Font(bold=True)
        ws_104.cell(row=row_num, column=1).border = border
        ws_104.cell(row=row_num, column=2).value = value
        ws_104.cell(row=row_num, column=2).number_format = '$#,##0.00'
        ws_104.cell(row=row_num, column=2).border = border
        ws_104.cell(row=row_num, column=2).alignment = Alignment(horizontal='right')
        
        if label == 'TOTAL PAGADO':
            ws_104.cell(row=row_num, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            ws_104.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF", size=12)
            ws_104.cell(row=row_num, column=2).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            ws_104.cell(row=row_num, column=2).font = Font(bold=True, size=12)
        
        row_num += 1
    
    # Column widths
    ws_104.column_dimensions['A'].width = 12
    ws_104.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws_104.column_dimensions[col].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"{razon_social.replace(' ', '_')}_{year}_resumen_anual.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ------------------------------
# Export to PDF
# ------------------------------
@router.post("/{razon_social}/export-pdf/{year}")
async def export_yearly_pdf(
    razon_social: str,
    year: str,
    branding: PDFBranding,
    exclude_months: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export yearly summary to branded PDF
    ✅ Professional layout with company branding
    ✅ User data isolation
    """
    if not is_valid_year(year):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid year parameter: '{year}'. Cannot export PDF for invalid year."
        )

    excluded = set(int(m.strip()) for m in exclude_months.split(',')) if exclude_months else set()

    # Fetch Form 103 documents
    query_103 = select(Document).where(
        and_(
            Document.razon_social == razon_social,
            Document.form_type == FormTypeEnum.FORM_103,
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    ).order_by(Document.periodo_mes_numero.asc())
    if excluded:
        query_103 = query_103.where(Document.periodo_mes_numero.notin_(excluded))
    docs_103 = (await db.execute(query_103)).scalars().all()

    # Fetch Form 104 documents
    query_104 = select(Document).where(
        and_(
            Document.razon_social == razon_social,
            Document.form_type == FormTypeEnum.FORM_104,
            Document.periodo_anio == year,
            Document.user_id == current_user.id
        )
    ).order_by(Document.periodo_mes_numero.asc())
    if excluded:
        query_104 = query_104.where(Document.periodo_mes_numero.notin_(excluded))
    docs_104 = (await db.execute(query_104)).scalars().all()

    # PDF setup
    output = BytesIO()
    pdf_doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor(branding.primary_color),
        spaceAfter=12,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor(branding.secondary_color),
        spaceAfter=10
    )

    # Logo
    if branding.logo_url:
        try:
            response = requests.get(branding.logo_url, timeout=5)
            img = PILImage.open(BytesIO(response.content))
            aspect = img.height / float(img.width)
            logo = Image(BytesIO(response.content), width=2*inch, height=(2*aspect)*inch)
            elements.append(logo)
            elements.append(Spacer(1, 0.2*inch))
        except:
            pass

    # Title
    elements.append(Paragraph(branding.company_name, title_style))
    elements.append(Paragraph(f"Resumen Anual {year}", title_style))
    elements.append(Paragraph(razon_social, subtitle_style))
    elements.append(Spacer(1, 0.3*inch))

    month_names = ['', 'Enero', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    # --- Form 103 Summary ---
    elements.append(Paragraph("Form 103 - Retenciones en la Fuente", subtitle_style))
    table_data_103 = [['Mes', 'Período', 'Subtotal Op.', 'Retención', 'Impuesto', 'Total Pagado']]

    total_subtotal = total_retencion = total_impuesto = total_pagado = 0

    for doc in docs_103:
        tot = (await db.execute(select(Form103Totals).where(Form103Totals.document_id == doc.id))).scalar_one_or_none()
        if not tot:
            continue
        periodo = f"{doc.periodo_mes[:3]} {year}" if doc.periodo_mes else "N/A"
        table_data_103.append([
            month_names[doc.periodo_mes_numero] if doc.periodo_mes_numero else "N/A",
            periodo,
            f"${tot.subtotal_operaciones_pais or 0:,.2f}",
            f"${tot.total_retencion or 0:,.2f}",
            f"${tot.total_impuesto_pagar or 0:,.2f}",
            f"${tot.total_pagado or 0:,.2f}"
        ])
        total_subtotal += tot.subtotal_operaciones_pais or 0.0
        total_retencion += tot.total_retencion or 0.0
        total_impuesto += tot.total_impuesto_pagar or 0.0
        total_pagado += tot.total_pagado or 0.0

    # Total row
    table_data_103.append([
        'TOTAL', '',
        f"${total_subtotal:,.2f}",
        f"${total_retencion:,.2f}",
        f"${total_impuesto:,.2f}",
        f"${total_pagado:,.2f}"
    ])

    table_103 = Table(table_data_103, colWidths=[0.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    table_103.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(branding.primary_color)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table_103)
    elements.append(Spacer(1, 0.4*inch))

    # --- Form 104 Summary ---
    elements.append(Paragraph("Form 104 - IVA", subtitle_style))
    table_data_104 = [['Mes', 'Ventas Neto', 'Imp. Gen.', 'Adquis.', 'Créd. Trib.', 'Total Pagado']]

    total_ventas = total_imp_gen = total_adq = total_cred = total_pag_104 = 0

    for doc in docs_104:
        data = (await db.execute(select(Form104Data).where(Form104Data.document_id == doc.id))).scalar_one_or_none()
        if not data:
            continue
        
        ventas = getattr(data, 'total_ventas_neto', 0) or 0
        imp_gen = getattr(data, 'total_impuesto_generado', 0) or 0
        adq = getattr(data, 'total_adquisiciones', 0) or 0
        cred = getattr(data, 'credito_tributario_aplicable', 0) or 0
        pag = getattr(data, 'total_pagado', 0) or 0
        
        table_data_104.append([
            month_names[doc.periodo_mes_numero] if doc.periodo_mes_numero else "N/A",
            f"${ventas:,.2f}",
            f"${imp_gen:,.2f}",
            f"${adq:,.2f}",
            f"${cred:,.2f}",
            f"${pag:,.2f}"
        ])
        total_ventas += ventas
        total_imp_gen += imp_gen
        total_adq += adq
        total_cred += cred
        total_pag_104 += pag

    # Total row
    table_data_104.append([
        'TOTAL',
        f"${total_ventas:,.2f}",
        f"${total_imp_gen:,.2f}",
        f"${total_adq:,.2f}",
        f"${total_cred:,.2f}",
        f"${total_pag_104:,.2f}"
    ])

    table_104 = Table(table_data_104, colWidths=[0.8*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    table_104.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(branding.secondary_color)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table_104)
    elements.append(Spacer(1, 0.4*inch))

    # Build the PDF
    pdf_doc.build(elements)
    output.seek(0)

    # Return as downloadable file
    filename = f"{razon_social.replace(' ', '_')}_{year}_resumen_anual.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )